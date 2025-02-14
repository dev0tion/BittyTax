# -*- coding: utf-8 -*-
# (c) Nano Nano Ltd 2019

import csv
import sys
import warnings
from decimal import Decimal, InvalidOperation

import dateutil.parser
import xlrd
from colorama import Back, Fore
from openpyxl import load_workbook
from tqdm import tqdm, trange

from .config import config
from .constants import ERROR, TZ_UTC
from .exceptions import (
    DataValueError,
    MissingDataError,
    TimestampParserError,
    TransactionParserError,
    UnexpectedDataError,
    UnexpectedTransactionTypeError,
)
from .record import TransactionRecord as TR
from .transactions import Buy, Sell


class ImportRecords:
    def __init__(self):
        self.t_rows = []
        self.success_cnt = 0
        self.failure_cnt = 0

    def import_excel_xlsx(self, filename):
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        workbook = load_workbook(filename=filename, read_only=False, data_only=True)
        print(f"{Fore.WHITE}Excel file: {Fore.YELLOW}{filename}")

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            if worksheet.title.startswith("--"):
                print(f"{Fore.GREEN}skipping '{worksheet.title}' worksheet")
                continue

            if config.debug:
                print(f"{Fore.CYAN}importing '{worksheet.title}' rows")

            for row_num, row in enumerate(
                tqdm(
                    worksheet.rows,
                    total=worksheet.max_row,
                    unit=" row",
                    desc=f"{Fore.CYAN}importing '{worksheet.title}' rows{Fore.GREEN}",
                    disable=bool(config.debug or not sys.stdout.isatty()),
                )
            ):
                if row_num == 0:
                    # Skip headers
                    continue

                row = [self.convert_cell_xlsx(cell) for cell in row]

                t_row = TransactionRow(
                    row[: len(TransactionRow.HEADER)], row_num + 1, worksheet.title
                )

                try:
                    t_row.parse()
                except TransactionParserError as e:
                    t_row.failure = e

                if config.debug or t_row.failure:
                    tqdm.write(f"{Fore.YELLOW}import: {t_row}")

                if t_row.failure:
                    tqdm.write(f"{ERROR} {t_row.failure}")

                self.t_rows.append(t_row)
                self.update_cnts(t_row)

        workbook.close()
        del workbook

    def import_excel_xls(self, filename):
        workbook = xlrd.open_workbook(filename)
        print(f"{Fore.WHITE}Excel file: {Fore.YELLOW}{filename}")

        for worksheet in workbook.sheets():
            if worksheet.name.startswith("--"):
                print(f"{Fore.GREEN}skipping '{worksheet.name}' worksheet")
                continue

            if config.debug:
                print(f"{Fore.CYAN}importing '{worksheet.name}' rows")

            for row_num in trange(
                0,
                worksheet.nrows,
                unit=" row",
                desc=f"{Fore.CYAN}importing '{worksheet.name}' rows{Fore.GREEN}",
                disable=bool(config.debug or not sys.stdout.isatty()),
            ):
                if row_num == 0:
                    # Skip headers
                    continue

                row = [
                    self.convert_cell_xls(worksheet.cell(row_num, cell_num), workbook)
                    for cell_num in range(0, worksheet.ncols)
                ]

                t_row = TransactionRow(
                    row[: len(TransactionRow.HEADER)], row_num + 1, worksheet.name
                )

                try:
                    t_row.parse()
                except TransactionParserError as e:
                    t_row.failure = e

                if config.debug or t_row.failure:
                    tqdm.write(f"{Fore.YELLOW}import: {t_row}")

                if t_row.failure:
                    tqdm.write(f"{ERROR} {t_row.failure}")

                self.t_rows.append(t_row)
                self.update_cnts(t_row)

        workbook.release_resources()
        del workbook

    @staticmethod
    def convert_cell_xlsx(cell):
        if cell.value is None:
            return ""
        return str(cell.value)

    @staticmethod
    def convert_cell_xls(cell, workbook):
        if cell.ctype == xlrd.XL_CELL_DATE:
            datetime = xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode)
            if datetime.microsecond:
                value = f"{datetime:%Y-%m-%dT%H:%M:%S.%f}"
            else:
                value = f"{datetime:%Y-%m-%dT%H:%M:%S}"
        elif cell.ctype in (
            xlrd.XL_CELL_NUMBER,
            xlrd.XL_CELL_BOOLEAN,
            xlrd.XL_CELL_ERROR,
        ):
            # repr is required to ensure no precision is lost
            value = repr(cell.value)
        else:
            value = str(cell.value)

        return value

    def import_csv(self, import_file):
        print(f"{Fore.WHITE}CSV file: {Fore.YELLOW}{import_file.name}")
        if config.debug:
            print(f"{Fore.CYAN}importing rows")

        reader = csv.reader(import_file)

        for row in tqdm(
            reader,
            unit=" row",
            desc=f"{Fore.CYAN}importing{Fore.GREEN}",
            disable=bool(config.debug or not sys.stdout.isatty()),
        ):
            if reader.line_num == 1:
                # Skip headers
                continue

            t_row = TransactionRow(row[: len(TransactionRow.HEADER)], reader.line_num)
            try:
                t_row.parse()
            except TransactionParserError as e:
                t_row.failure = e

            if config.debug or t_row.failure:
                tqdm.write(f"{Fore.YELLOW}import: {t_row}")

            if t_row.failure:
                tqdm.write(f"{ERROR} {t_row.failure}")

            self.t_rows.append(t_row)
            self.update_cnts(t_row)

    def update_cnts(self, t_row):
        if t_row.failure is not None:
            self.failure_cnt += 1
        elif t_row.t_record is not None:
            self.success_cnt += 1

    def get_records(self):
        transaction_records = [t_row.t_record for t_row in self.t_rows if t_row.t_record]

        transaction_records.sort()
        for t_record in transaction_records:
            t_record.set_tid()

        if config.debug:
            for t_row in self.t_rows:
                print(f"{Fore.YELLOW}import: {t_row}")

        return transaction_records


class TransactionRow:
    HEADER = [
        "Type",
        "Buy Quantity",
        "Buy Asset",
        "Buy Value",
        "Sell Quantity",
        "Sell Asset",
        "Sell Value",
        "Fee Quantity",
        "Fee Asset",
        "Fee Value",
        "Wallet",
        "Timestamp",
        "Note",
    ]

    OPT = "Optional"
    MAN = "Mandatory"

    TYPE_VALIDATION = {
        TR.TYPE_DEPOSIT: [MAN, MAN, MAN, OPT, None, None, None, OPT, OPT, OPT],
        TR.TYPE_MINING: [MAN, MAN, MAN, OPT, None, None, None, OPT, OPT, OPT],
        TR.TYPE_STAKING: [MAN, MAN, MAN, OPT, None, None, None, OPT, OPT, OPT],
        TR.TYPE_INTEREST: [MAN, MAN, MAN, OPT, None, None, None, OPT, OPT, OPT],
        TR.TYPE_DIVIDEND: [MAN, MAN, MAN, OPT, None, None, None, OPT, OPT, OPT],
        TR.TYPE_INCOME: [MAN, MAN, MAN, OPT, None, None, None, OPT, OPT, OPT],
        TR.TYPE_GIFT_RECEIVED: [MAN, MAN, MAN, OPT, None, None, None, OPT, OPT, OPT],
        TR.TYPE_AIRDROP: [MAN, MAN, MAN, OPT, None, None, None, OPT, OPT, OPT],
        TR.TYPE_WITHDRAWAL: [MAN, None, None, None, MAN, MAN, OPT, OPT, OPT, OPT],
        TR.TYPE_SPEND: [MAN, None, None, None, MAN, MAN, OPT, OPT, OPT, OPT],
        TR.TYPE_GIFT_SENT: [MAN, None, None, None, MAN, MAN, OPT, OPT, OPT, OPT],
        TR.TYPE_GIFT_SPOUSE: [MAN, None, None, None, MAN, MAN, OPT, OPT, OPT, OPT],
        TR.TYPE_CHARITY_SENT: [MAN, None, None, None, MAN, MAN, OPT, OPT, OPT, OPT],
        TR.TYPE_LOST: [MAN, None, None, None, MAN, MAN, OPT, None, None, None],
        TR.TYPE_TRADE: [MAN, MAN, MAN, OPT, MAN, MAN, OPT, OPT, OPT, OPT],
    }

    TRANSFER_TYPES = (TR.TYPE_DEPOSIT, TR.TYPE_WITHDRAWAL)

    def __init__(self, row, row_num, worksheet_name=None):
        self.row = row
        self.row_dict = dict(zip(self.HEADER, row))
        self.row_num = row_num
        self.worksheet_name = worksheet_name
        self.t_record = None
        self.failure = None

    def parse(self):
        if all(not self.row[i] for i in range(len(self.row) - 1)):
            # Skip empty rows
            return

        buy = sell = fee = None
        t_type = self.row_dict["Type"]

        if t_type not in self.TYPE_VALIDATION:
            raise UnexpectedTransactionTypeError(self.HEADER.index("Type"), "Type", t_type)

        for pos, required in enumerate(self.TYPE_VALIDATION[t_type]):
            if pos == self.HEADER.index("Buy Quantity"):
                buy_quantity = self.validate_quantity("Buy Quantity", required)
            elif pos == self.HEADER.index("Buy Asset"):
                buy_asset = self.validate_asset("Buy Asset", required)
            elif pos == self.HEADER.index("Buy Value"):
                buy_value = self.validate_value("Buy Value", required)
            elif pos == self.HEADER.index("Sell Quantity"):
                sell_quantity = self.validate_quantity("Sell Quantity", required)
            elif pos == self.HEADER.index("Sell Asset"):
                sell_asset = self.validate_asset("Sell Asset", required)
            elif pos == self.HEADER.index("Sell Value"):
                sell_value = self.validate_value("Sell Value", required)
            elif pos == self.HEADER.index("Fee Quantity"):
                fee_quantity = self.validate_quantity("Fee Quantity", required)
            elif pos == self.HEADER.index("Fee Asset"):
                fee_asset = self.validate_asset("Fee Asset", required)
            elif pos == self.HEADER.index("Fee Value"):
                fee_value = self.validate_value("Fee Value", required)

        if buy_value and buy_asset == config.ccy and buy_value != buy_quantity:
            raise DataValueError(self.HEADER.index("Buy Value"), "Buy Value", buy_value)

        if sell_value and sell_asset == config.ccy and sell_value != sell_quantity:
            raise DataValueError(self.HEADER.index("Sell Value"), "Sell Value", sell_value)

        if fee_value and fee_asset == config.ccy and fee_value != fee_quantity:
            raise DataValueError(self.HEADER.index("Fee Value"), "Fee Value", fee_value)

        if fee_quantity is not None and not fee_asset:
            raise MissingDataError(self.HEADER.index("Fee Asset"), "Fee Asset")

        if fee_quantity is None and fee_asset:
            raise MissingDataError(self.HEADER.index("Fee Quantity"), "Fee Quantity")

        if buy_asset:
            buy = Buy(t_type, buy_quantity, buy_asset, buy_value)
        if sell_asset:
            if t_type == TR.TYPE_LOST:
                if sell_value is None:
                    sell_value = Decimal(0)

                sell = Sell(t_type, sell_quantity, sell_asset, sell_value)
                if config.lost_buyback:
                    buy = Buy(t_type, sell_quantity, sell_asset, sell_value)
                    buy.acquisition = True
            else:
                sell = Sell(t_type, sell_quantity, sell_asset, sell_value)
        if fee_asset:
            # Fees are added as a separate spend transaction
            fee = Sell(TR.TYPE_SPEND, fee_quantity, fee_asset, fee_value)

            # Transfers fees are a special case
            if t_type in self.TRANSFER_TYPES:
                if config.transfers_include:
                    # Not a disposal, fees removed from the pool at zero cost
                    fee.disposal = False
                else:
                    # Not a disposal (unless configured otherwise)
                    if not config.transfer_fee_disposal:
                        fee.disposal = False

        if len(self.row) == len(self.HEADER):
            note = self.row_dict["Note"]
        else:
            note = ""

        self.t_record = TR(
            t_type,
            buy,
            sell,
            fee,
            self.row_dict["Wallet"],
            self.parse_timestamp(),
            note,
        )

    def parse_timestamp(self):
        try:
            timestamp = dateutil.parser.parse(self.row_dict["Timestamp"])
        except ValueError as e:
            raise TimestampParserError(
                self.HEADER.index("Timestamp"), "Timestamp", self.row_dict["Timestamp"]
            ) from e

        if timestamp.tzinfo is None:
            # Default to UTC if no timezone is specified
            timestamp = timestamp.replace(tzinfo=TZ_UTC)

        return timestamp

    def validate_quantity(self, quantity_hdr, required):
        if self.row_dict[quantity_hdr]:
            if required:
                try:
                    quantity = Decimal(self.strip_non_digits(self.row_dict[quantity_hdr]))
                except InvalidOperation as e:
                    raise DataValueError(
                        self.HEADER.index(quantity_hdr),
                        quantity_hdr,
                        self.row_dict[quantity_hdr],
                    ) from e

                if quantity < 0:
                    raise DataValueError(self.HEADER.index(quantity_hdr), quantity_hdr, quantity)
                return quantity

            raise UnexpectedDataError(
                self.HEADER.index(quantity_hdr),
                quantity_hdr,
                self.row_dict[quantity_hdr],
            )
        if required == self.MAN:
            raise MissingDataError(self.HEADER.index(quantity_hdr), quantity_hdr)

        return None

    def validate_asset(self, asset_hdr, required):
        if self.row_dict[asset_hdr]:
            if required:
                return self.row_dict[asset_hdr]

            raise UnexpectedDataError(
                self.HEADER.index(asset_hdr), asset_hdr, self.row_dict[asset_hdr]
            )
        if required == self.MAN:
            raise MissingDataError(self.HEADER.index(asset_hdr), asset_hdr)

        return None

    def validate_value(self, value_hdr, required):
        if self.row_dict[value_hdr]:
            if required:
                try:
                    value = Decimal(self.strip_non_digits(self.row_dict[value_hdr]))
                except InvalidOperation as e:
                    raise DataValueError(
                        self.HEADER.index(value_hdr),
                        value_hdr,
                        self.row_dict[value_hdr],
                    ) from e

                if value < 0:
                    raise DataValueError(self.HEADER.index(value_hdr), value_hdr, value)

                return value

            raise UnexpectedDataError(
                self.HEADER.index(value_hdr), value_hdr, self.row_dict[value_hdr]
            )

        if required == self.MAN:
            raise MissingDataError(self.HEADER.index(value_hdr), value_hdr)

        return None

    @staticmethod
    def strip_non_digits(string):
        return string.strip("£€$").replace(",", "")

    def __str__(self):
        if self.t_record and self.t_record.tid:
            tid_str = f" {Fore.MAGENTA}[TID:{self.t_record.tid[0]}]"
        else:
            tid_str = ""

        if self.worksheet_name:
            worksheet_str = f"'{self.worksheet_name}' "
        else:
            worksheet_str = ""

        row_str = ", ".join(
            [
                f"{Back.RED}'{data}'{Back.RESET}"
                if self.failure and self.failure.col_num == num
                else f"'{data}'"
                for num, data in enumerate(self.row)
            ]
        )

        return f"{worksheet_str}row[{self.row_num}] [{row_str}]{tid_str}"
